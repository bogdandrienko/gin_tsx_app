import collections
import datetime
import random
import re
import time
from typing import Union

import cx_Oracle
import openpyxl
from django.conf import settings
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import Group, User, update_last_login
from django.core.cache import caches
from django.db import connection, transaction
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import render
from django.views import View
from django.views.decorators.cache import cache_page
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from django_app import models as django_models
from django_app import serializers as django_serializers
from django_app import signals as django_signals
from django_app import utils as django_utils
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

django_signals.register_all_signals()
http_method_names = ["GET", "POST", "PUT", "PATCH", "DELETE", "HEAD", "OPTIONS"]

LocMemCache = caches["default"]
DatabaseCache = caches["special"]


# RedisCache = caches["extra"]


# Create your views here.

def predictive_f(request):
    try:
        context = {}
        return render(request, "build/index.html", context=context)
    except Exception as error:
        if settings.DEBUG:
            print(f"error {error}")
        return django_utils.DjangoClass.DRFClass.RequestOldClass.return_global_error(request=request, error=error)


@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=False)
def vehtrips_status_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> dict | str | None:
    if request.method == "GET":
        workbook: Workbook = openpyxl.load_workbook("./PredictiveAnalytic/src/vehtrips.xlsx")
        worksheet: Worksheet = workbook.active
        __data: tuple[tuple[Cell]] = worksheet.iter_rows(
            min_col=worksheet.min_column, max_col=worksheet.max_column,
            min_row=worksheet.min_row, max_row=worksheet.max_row,
            values_only=True
        )
        # result = [tuple([j.value if j.value is not None else "" for j in i if j.value]) for i in __data]
        result = [tuple([j if j is not None else "" for j in i]) for i in __data]
        rows_instances = [
            {"vehid": i[0], "time": i[1], "x": i[2], "y": i[3], "weight": i[4], "fuel": i[5], "speed": i[6]}
            for i in result]
        return rows_instances[2:]
    if request.method == "GET1":
        try:
            cx_Oracle.init_oracle_client(lib_dir=r"static_external/instantclient_21_9_lite")
        except Exception as error:
            pass

        with cx_Oracle.connect('DISPATCHER/disp@172.30.23.16/PITENEW') as __connection:
            with __connection.cursor() as __cursor:
                query = """
SELECT VEHID, 
       TIME, 
       X, 
       Y, 
       WEIGHT, 
       FUEL, 
       SPEED 
FROM   EVENTSTATEARCHIVE 
WHERE  MESCOUNTER IN (SELECT MAX(MESCOUNTER) 
                      FROM   EVENTSTATEARCHIVE 
                      WHERE  TIME BETWEEN ( SYSDATE - ( 1 / 24 / 60 * 500 ) ) AND SYSDATE 
                      GROUP  BY VEHID) 
ORDER  BY VEHID 
"""
                __cursor.execute(query)
                rows_raw = __cursor.fetchall()
                rows_instances = [
                    {"vehid": i[0], "time": i[1], "x": i[2], "y": i[3], "weight": i[4], "fuel": i[5], "speed": i[6]}
                    for i in rows_raw]

        # time.sleep(random.randint(1, 3))
        return {"data": rows_instances}


# @cache_page(timeout=1)
def index_f(request):
    try:
        context = {}
        return render(request, "build/index.html", context=context)
    except Exception as error:
        if settings.DEBUG:
            print(f"error {error}")
        return django_utils.DjangoClass.DRFClass.RequestOldClass.return_global_error(request=request, error=error)


@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=False)
def captcha_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> dict | str | None:
    if request.method == "GET":
        time.sleep(random.randint(1, 3))
        return "Вы не робот!"


@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=False)
def token_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> dict | str | None:
    if request.method == "POST":
        # TODO {"username": "admin", "password": "admin"}
        # TODO {"username": "user", "password": "12345Qwerty!"}
        username_str = request.get(key="username", _type=str, default="", is_file=False)
        password_str = request.get(key="password", _type=str, default="", is_file=False)

        if username_str and password_str:
            user_obj = authenticate(username=username_str, password=password_str)
            if user_obj is None:
                raise Exception("Username or password incorrect!")
            if user_obj.user_model.is_active_account is False:
                raise Exception("Attention, your account is banned!")
            access_count = 0
            for log in django_models.LoggingModel.objects.filter(
                    username=User.objects.get(id=username_str),
                    ip=request.ip,
                    path=request.path,
                    method=f"{request.method} | {request.action}",
                    error="-",
            ):
                if (log.created + datetime.timedelta(hours=6, minutes=59)).strftime("%Y-%m-%d %H:%M") >= (
                        datetime.datetime.now()
                ).strftime("%Y-%m-%d %H:%M"):
                    access_count += 1
            if access_count < 20:
                update_last_login(sender=None, user=user_obj)
                token_str = django_models.TokenModel.create_or_update_token(user_obj=user_obj)
                return {"token": token_str}
            else:
                raise Exception("Too many try for login!")
        raise Exception("Username or password not fulled!")


@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=True)
def detail_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> dict | str | None:
    if request.method == "GET":
        user_obj = django_utils.DjangoClass.Caching.cache(
            key=f"user_obj {request.user.username}",
            lambda_queryset=lambda: django_serializers.UserSerializer(request.user, many=False).data,
            cache_instance=LocMemCache,
            timeout=10,
        )

        user_model_obj = django_utils.DjangoClass.Caching.cache(
            key=f"user_model_obj {request.user.username}",
            lambda_queryset=lambda: django_serializers.UserModelSerializer(request.user_model, many=False).data,
            cache_instance=LocMemCache,
            timeout=10,
        )
        return {"user": user_obj, "user_model": user_model_obj}


@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=False)
def todo_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> dict | str | None:
    if request.pk:
        if request.method == "GET":
            if request.action == "_":
                todo_obj = django_models.Todo.objects.get(id=request.pk)
                return django_serializers.TodoSerializer(todo_obj, many=False).data
        elif request.method == "PUT" or request.method == "PATCH":
            if request.action == "_":
                title = request.get(key="title", _type=str, default=None, is_file=False)
                description = request.get(key="description", _type=str, default=None, is_file=False)
                avatar = request.get(key="avatar", _type=any, default=None, is_file=True)
                is_completed = request.get(key="is_completed", _type=bool, default=None, is_file=True)

                todo_obj = django_models.Todo.objects.get(id=request.pk)
                if title is not None and todo_obj.title != title:
                    todo_obj.title = title
                if description is not None and todo_obj.description != description:
                    todo_obj.description = description
                if avatar is not None and todo_obj.avatar != avatar:
                    todo_obj.avatar = avatar
                if is_completed is not None and todo_obj.is_completed != is_completed:
                    todo_obj.is_completed = is_completed
                todo_obj.save()
                return "Successfully update"
        elif request.method == "DELETE":
            if request.action == "":
                django_models.Todo.objects.get(id=request.pk).delete()
                return "Successfully delete"
    else:
        if request.method == "GET":
            if request.action == "_":
                page = request.get(key="page", _type=int, default=None, is_file=False)
                limit = request.get(key="limit", _type=int, default=None, is_file=False)

                todos_obj = django_models.Todo.objects.all()
                if page and limit:
                    todos_obj = django_utils.DjangoClass.PaginationClass.paginate(
                        page_number=page, object_list=todos_obj, limit_per_page=limit
                    )
                return django_serializers.TodoSerializer(todos_obj, many=True).data
        elif request.method == "POST":
            if request.action == "_":
                title = request.get(key="title", _type=str, default=None, is_file=False)
                description = request.get(key="description", _type=str, default=None, is_file=False)
                avatar = request.get(key="avatar", _type=any, default=None, is_file=True)
                is_completed = request.get(key="is_completed", _type=bool, default=False, is_file=False)

                if title:
                    django_models.Todo.objects.create(
                        title=title, description=description, avatar=avatar, is_completed=is_completed
                    )
                    return "Successfully create"
                else:
                    raise Exception({"detail": "Not have data"})


@django_utils.DjangoClass.DRFClass.RequestClass.request(auth=False)
def report_f(request: django_utils.DjangoClass.DRFClass.RequestClass) -> dict | str | None:
    if request.pk:
        pass
    else:
        if request.method == "GET":
            if request.action == "monitoring":
                data = [
                    {
                        "id": x,
                        "type": "Автосамосвал",
                        "speed": random.randint(0, 20),
                        "mass": random.randint(87, 102),
                        "status": random.choice(["Норма", "Простой", "Движение", "Погрузка", "Ремонт", "ППР"]),
                        "time": django_utils.DateTimeUtils.get_current_time(),
                    }
                    for x in range(201, 230)
                ]
                return {"data": data}
            elif request.action == "report":
                try:
                    # ?page=1&limit=10&sort_by=name%20(asc)&filter_by=pay%20(true)&search=95
                    page = request.GET.get("page", 1)
                    limit = request.GET.get("limit", 10)
                    search = request.GET.get("search", "")
                    filter_by = request.GET.get("filter_by", "")
                    sort_by = request.GET.get("sort_by", "")

                    workbook = openpyxl.load_workbook("static/media/vehtrips.xlsx")
                    worksheet = workbook.active
                    matrix = worksheet.iter_rows(
                        min_col=1, min_row=1, max_col=worksheet.max_column, max_row=300, values_only=True
                    )
                    # matrix = []
                    # for i in range(1, worksheet.max_row+1):
                    #     row_data = []
                    #     for j in range(1, worksheet.max_column+1):
                    #         print(i, j)
                    #         row_data.append(worksheet.cell(i, j).value)
                    #     matrix.append(row_data)
                    matrix = [
                        {
                            "Автосамосвал": x[0],
                            "Экскаватор": x[1],
                            "Зона разгрузки": x[2],
                            "Тип породы": x[3],
                            "Зона погрузки": x[4],
                            "Ном.": x[5],
                            "Расстояние": x[6],
                            "Масса": x[7],
                            "Объём": x[8],
                            "Частота по массе": x[9],
                            "Частота по объёму": x[10],
                            "Сред. скорость": x[11],
                            "Часы": x[12],
                            "Высота погрузки": x[13],
                            "Высота разгрузки": x[14],
                            "Время погрузки": x[15],
                            "Время выезда на дорогу": x[16],
                            "Время разгрузки": x[17],
                            "Время рейса": x[18],
                        }
                        for x in list(matrix)
                    ]
                    matrix = [
                        x
                        for x in matrix
                        if isinstance(x["Время разгрузки"], datetime.datetime)
                           and (filter_by == "все" or filter_by == x["Тип породы"])
                           and len(x["Зона разгрузки"]) > 0
                    ]
                    if sort_by == "времени погрузки (свежие в начале)":
                        matrix.sort(key=lambda x: x["Время погрузки"], reverse=True)
                    else:
                        matrix.sort(key=lambda x: x["Время погрузки"])
                    return {"data": matrix}
                except Exception as error:
                    print(error)
                    return {"data": []}
