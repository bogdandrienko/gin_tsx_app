import datetime
import sys
import threading
import time
import cx_Oracle
import openpyxl
from PyQt6 import QtCore, QtGui, QtWidgets
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# https://habr.com/ru/company/skillfactory/blog/599599/
# https://www.pythonguis.com/pyqt6-tutorial/
# https://pythonworld.ru/gui/pyqt5-firstprograms.html

# call env/scripts/activate
# pip install PyQt6 auto-py-to-exe cx_Oracle openpyxl
# pip install -r requirements.txt
# pip freeze > requirements.txt
# import pyinstaller  # pyinstaller --noconfirm --windowed --onedir predictive.py
# auto-py-to-exe

class Queries(object):
    @staticmethod
    def get_excel_all_trips() -> list[tuple]:
        workbook: Workbook = openpyxl.load_workbook("./src/vehtrips.xlsx")
        worksheet: Worksheet = workbook.active
        __data: tuple[tuple[Cell]] = worksheet.iter_rows(
            min_col=worksheet.min_column, max_col=worksheet.max_column,
            min_row=worksheet.min_row, max_row=worksheet.max_row,
            values_only=True
        )
        # result = [tuple([j.value if j.value is not None else "" for j in i if j.value]) for i in __data]
        result = [tuple([j if j is not None else "" for j in i]) for i in __data]
        # print(result)
        # for i in result:
        #     print(i)
        return result

    @staticmethod
    def get_connection_conf() -> str:
        return 'DISPATCHER/disp@172.30.23.16/PITENEW'

    @staticmethod
    def execute_many(query_string: str) -> list[tuple[any]]:
        with cx_Oracle.connect(Queries.get_connection_conf()) as __connection:
            with __connection.cursor() as __cursor:
                __cursor.execute(query_string)
                return __cursor.fetchall()

    @staticmethod
    def get_last_trip() -> str:
        return """
SELECT * 
FROM   (SELECT VEHID, 
               SHOVID, 
               UNLOADID, 
               WORKTYPE, 
               TIMELOAD, 
               TIMEUNLOAD, 
               TIME_INSERTING, 
               MOVETIME, 
               round((TIMEUNLOAD - TIMELOAD + (MOVETIME- TO_DATE('2000/01/01', 'yyyy/mm/dd')) / 1.65) * 24 * 60 , 1) TRIPTIME, 
               AREA, 
               FUELLOAD, 
               FUELUNLOAD, 
               WEIGHT, 
               AVSPEED, 
               LENGTH, 
               UNLOADLENGTH, 
               1 AS TripNum
        FROM VEHTRIPS vt
      WHERE  vt.TIMEUNLOAD BETWEEN GETPREDEFINEDTIMEFROM('за указанную смену', GETCURSHIFTNUM(0, SYSDATE)-1, GETCURSHIFTDATE(0, SYSDATE)) AND GETPREDEFINEDTIMETO('за указанную смену', GETCURSHIFTNUM(0, SYSDATE), GETCURSHIFTDATE(0, SYSDATE))
               AND VEHID = '139'
        ORDER  BY TIMELOAD DESC) 
WHERE ROWNUM < 2
"""

    @staticmethod
    def get_last_trips_new() -> str:
        return """
    SELECT 
    tripcounter,
    VEHID, 
                   SHOVID, 
                   UNLOADID, 
                   WORKTYPE, 
                   
                   TIMELOAD, 
                   TIMEUNLOAD, 
                   MOVETIME, 
                   
                   WEIGHT, 
                   bucketcount, 
                   AVSPEED, 
                   LENGTH, 
                   UNLOADLENGTH, 
                   LOADHEIGHT, 
                   UNLOADHEIGHT
            FROM VEHTRIPS vt
          WHERE  vt.TIMEUNLOAD BETWEEN GETPREDEFINEDTIMEFROM('за указанную смену', GETCURSHIFTNUM(0, SYSDATE), GETCURSHIFTDATE(0, SYSDATE)) AND GETPREDEFINEDTIMETO('за указанную смену', GETCURSHIFTNUM(0, SYSDATE), GETCURSHIFTDATE(0, SYSDATE))
            ORDER  BY TIMELOAD DESC
    """


class TripInstance(object):
    def __init__(
            self,
            # todo идентификатор рейса
            trip_id: int,  # id рейса
            vehid: str,  # самосвал
            shovid: str,  # экскаватор
            place_unload: str,  # место разгрузки
            material: str,  # материал

            # todo временные показатели рейса
            time_load: datetime.datetime,  # время погрузки
            time_unload: datetime.datetime,  # время разгрузки
            time_move: datetime.datetime,  # время движения

            # todo показатели рейса
            weight: float,  # вес
            buckets_count: int,  # ковши
            speed: float,  # скорость
            length_after_load: float,  # расстояние после погрузки
            length_after_unload: float,  # расстояние после разгрузки
            height_load: int,  # высота погрузки
            height_unload: int,  # высота разгрузки
    ):
        # todo идентификатор рейса
        self.trip_id = trip_id
        self.vehid = vehid
        self.shovid = shovid
        self.place_unload = place_unload
        self.material = material

        # todo временные показатели рейса
        self.time_load = time_load
        self.time_unload = time_unload
        self.time_move = time_move

        # todo показатели рейса
        self.weight = weight
        self.buckets_count = buckets_count
        self.speed = speed
        self.length_after_load = length_after_load
        self.length_after_unload = length_after_unload
        self.height_load = height_load
        self.height_unload = height_unload

    def get_norm_time_full(
            self, norm_shovels: dict[str, dict[str, any]], norm_dumptrucks: dict[str, dict[str, any]]
    ) -> str:
        time_load = round(norm_shovels[self.shovid][self.material], 2)
        time_path = round(self.length_after_load / norm_dumptrucks[self.vehid]["Ср. скорость гружённый"] * 60, 2)
        time_unload = round(norm_dumptrucks[self.vehid]["Время разгрузки"], 2)
        time_return = round(self.length_after_load / norm_dumptrucks[self.vehid]["Ср. скорость порожний"] * 60, 2)
        time_wait = round(norm_dumptrucks[self.vehid]["Время ожидания"], 2)

        time_norm: float = round(time_load + time_path + time_unload + time_return + time_wait, 2)
        time_fact: float = round((self.time_unload - self.time_load).total_seconds() / 60 + time_return + time_wait, 2)

        full_res = f"погрузка({time_load}) + путь({time_path}) + разгрузка({time_unload}) + возврат({time_return}) + " \
                   f"ожидание({time_wait}) = итого норма({time_norm}) / итого факт ({time_fact})"
        short_res = round(time_norm / time_fact * 100, 1)
        # print(full_res)
        return full_res + "\t\t" + str(short_res)

    class Norm:
        @staticmethod
        def get_norm_shovels() -> dict[str, dict[str, any]]:
            __workbook: Workbook = openpyxl.load_workbook("./src/table_norms.xlsx")
            __worksheet: Worksheet = __workbook["Экскаваторы"]
            __norms: dict[str, dict[str, any]] = {}
            for __i in range(__worksheet.min_row + 1, __worksheet.max_row + 1):
                __norms[str(__worksheet.cell(row=__i, column=1).value)] = {
                    "Вскрыша скальная": float(__worksheet.cell(row=__i, column=2).value),
                    "Вскрыша рыхлая": float(__worksheet.cell(row=__i, column=3).value),
                    "Вскрыша транзитная": float(__worksheet.cell(row=__i, column=4).value),
                    "Руда скальная": float(__worksheet.cell(row=__i, column=5).value),
                    "ВКП скала": float(__worksheet.cell(row=__i, column=6).value),
                }
            return __norms

        @staticmethod
        def get_norm_dumptrucks() -> dict[str, dict[str, any]]:
            __workbook: Workbook = openpyxl.load_workbook("./src/table_norms.xlsx")
            __worksheet: Worksheet = __workbook["Самосвалы"]
            __norms: dict[str, dict[str, any]] = {}
            for __i in range(__worksheet.min_row + 1, __worksheet.max_row + 1):
                __norms[str(__worksheet.cell(row=__i, column=1).value)] = {
                    "Состояние": float(__worksheet.cell(row=__i, column=2).value),
                    "Ср. скорость гружённый": float(__worksheet.cell(row=__i, column=3).value),
                    "Ср. скорость порожний": float(__worksheet.cell(row=__i, column=4).value),
                    "Время разгрузки": float(__worksheet.cell(row=__i, column=5).value),
                    "Время ожидания": float(__worksheet.cell(row=__i, column=6).value),
                }
            return __norms


class TripParser(object):
    @staticmethod
    def parse_from_tuple(_row: tuple) -> TripInstance:  # self
        return TripInstance(
            # todo идентификатор рейса
            trip_id=int(_row[0]),  # id рейса
            vehid=str(_row[1]),  # самосвал
            shovid=str(_row[2]),  # экскаватор
            place_unload=str(_row[3]),  # место разгрузки
            material=str(_row[4]),  # материал

            # todo временные показатели рейса
            time_load=_row[5],  # время погрузки
            time_unload=_row[6],  # время разгрузки
            time_move=_row[7],  # время движения

            # todo показатели рейса
            weight=float(_row[8]),  # вес
            buckets_count=int(_row[9] if len(str(_row[9])) > 0 and _row[9] is not None else 0),  # ковши
            speed=float(_row[10]),  # скорость
            length_after_load=float(_row[11]),  # расстояние после погрузки
            length_after_unload=float(_row[12]),  # расстояние после разгрузки
            height_load=int(_row[13]),  # высота погрузки
            height_unload=int(_row[14]),  # высота разгрузки
        )

    @staticmethod
    def parse_from_tuples(_rows: list[tuple]) -> list[TripInstance]:  # self
        return [TripParser.parse_from_tuple(_row=__i) for __i in _rows]


class Ui(object):
    class MainWindow(QtWidgets.QMainWindow):
        def __init__(self):
            super().__init__()

            self.pause = False

            self.setWindowTitle("Предиктивная аналитика")
            self.icon = QtGui.QIcon('src/images/icon3.png')
            self.setWindowIcon(self.icon)
            self.setGeometry(0, 0, 1280, 720)
            self.setMinimumSize(QtCore.QSize(640, 480))
            self.setMaximumSize(QtCore.QSize(1900, 1000))
            self.layout = QtWidgets.QGridLayout()
            widget = QtWidgets.QWidget()
            widget.setLayout(self.layout)

            self.button_start = QtWidgets.QPushButton("start")
            self.button_start.clicked.connect(self.start)
            self.layout.addWidget(self.button_start, 0, 2)

            self.button_stop = QtWidgets.QPushButton("stop")
            self.button_stop.clicked.connect(self.stop)
            self.layout.addWidget(self.button_stop, 0, 3)

            self.label_text = QtWidgets.QLabel("Запустите анализ")
            self.layout.addWidget(self.label_text, 1, 1)

            label = QtWidgets.QLabel(self)
            pixmap = QtGui.QPixmap('src/images/dumptruck2.png')
            pixmap = pixmap.scaled(
                QtCore.QSize(512, 512),
            )
            label.setPixmap(pixmap)
            self.layout.addWidget(label, 2, 3)

            self.vbox = QtWidgets.QVBoxLayout()
            for i in range(1, 50):
                self.vbox.addWidget(QtWidgets.QLabel("TextLabel"))

            self.widget = QtWidgets.QWidget()
            self.widget.setLayout(self.vbox)

            self.scroll = QtWidgets.QScrollArea()
            self.scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
            self.scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
            self.scroll.setWidgetResizable(True)
            self.scroll.setWidget(self.widget)
            self.layout.addWidget(self.scroll, 3, 4)

            self.tray__icon = QtWidgets.QSystemTrayIcon(self)
            self.tray__icon.setIcon(self.icon)

            '''
                Объявим и добавим действия для работы с иконкой системного трея
                show - показать окно
                hide - скрыть окно
                exit - выход из программы
            '''
            show__action = QtGui.QAction("Show", self)
            quit__action = QtGui.QAction("Exit", self)
            hide__action = QtGui.QAction("Hide", self)
            show__action.triggered.connect(self.show)
            hide__action.triggered.connect(self.hide)
            quit__action.triggered.connect(QtWidgets.QApplication.quit)
            quit__action.triggered.connect(self.closeEvent)
            tray__menu = QtWidgets.QMenu()
            tray__menu.addAction(show__action)
            tray__menu.addAction(hide__action)
            tray__menu.addAction(quit__action)
            self.tray__icon.setContextMenu(tray__menu)
            self.tray__icon.show()

            self.setCentralWidget(widget)

        def closeEvent(self, event):
            event.ignore()
            self.hide()
            self.tray__icon.showMessage(
                "Tray Program",
                "Application was minimized to Tray",
                QtWidgets.QSystemTrayIcon.Information,
                2000
            )

        def start(self):
            self.pause = False
            time.sleep(1.5)
            thread_2 = threading.Thread(target=self.click, args=(), kwargs={})
            thread_2.start()

        def stop(self):
            self.pause = True

        def click(self):
            print("started")
            while not self.pause:
                norm_shovels = TripInstance.Norm.get_norm_shovels()
                norm_dumptrucks = TripInstance.Norm.get_norm_dumptrucks()

                # todo get trip
                trips_raw: list[tuple] = Queries.execute_many(query_string=Queries.get_last_trips_new())[:10]
                # data: list[tuple] = Queries.get_excel_all_trips()[-5:]
                # todo get trip

                # todo extract trip
                trips_instances: list[TripInstance] = TripParser.parse_from_tuples(_rows=trips_raw)
                # todo extract trip

                # todo show result
                txts = f"Результаты ({datetime.datetime.now().strftime('%H:%M:%S')}): \n\n"
                for i in trips_instances:
                    txt = f'{i.vehid}({i.shovid})[{i.material}]|{i.time_load.strftime("%H:%M")}|:\t\t' \
                          f'{i.get_norm_time_full(norm_shovels=norm_shovels, norm_dumptrucks=norm_dumptrucks)}%\n'
                    txts += txt
                self.label_text.setText(txts)
                # todo show result

                time.sleep(2.0)
            print("stopped")
            pass

    @staticmethod
    def start():
        app = QtWidgets.QApplication(sys.argv)
        app.setStyle('Fusion')

        window = Ui.MainWindow()
        window.show()

        app.exec()


if __name__ == "__main__":
    try:
        cx_Oracle.init_oracle_client(lib_dir=r"src/instantclient_21_9_lite")

        # todo ui
        thread_1 = threading.Thread(target=Ui.start, args=(), kwargs={})
        thread_1.start()
        thread_1.join()
        # todo ui

        pass
    except Exception as error:
        print(error)
        input("\npress any key to exit\n")
